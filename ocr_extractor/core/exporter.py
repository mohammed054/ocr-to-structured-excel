"""Excel and JSON export logic for parsed OCR documents."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

try:
    from .parser import ParsedDocument
except ImportError:  # pragma: no cover - supports direct script execution
    from parser import ParsedDocument


EXCEL_SUFFIX = ".xlsx"
JSON_SUFFIX = ".json"
OUTPUT_STEM_SUFFIX = "_extracted"
MASTER_WORKBOOK_PREFIX = "master_extraction"
SUMMARY_SHEET_NAME = "Summary"
KEY_VALUES_SHEET_NAME = "Key_Values"
NO_TABLES_SHEET_NAME = "No_Tables"
MASTER_SUMMARY_SHEET_NAME = "Master_Summary"
MAX_SHEET_NAME_LENGTH = 31
HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
ALTERNATE_ROW_FILL = "EAF2F8"
SECTION_FILL = "D9EAF7"
WHITE_FILL = "FFFFFF"
DEFAULT_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 60
MIN_COLUMN_WIDTH = 10
INVALID_SHEET_CHARS = re.compile(r"[\[\]\*\?/\\:]")
INVALID_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9_.-]+")


@dataclass(frozen=True)
class ExportResult:
    """Paths and counts produced by an export operation."""

    input_file: str
    excel_path: Path
    json_path: Path
    table_count: int
    key_value_count: int
    page_count: int


@dataclass(frozen=True)
class BulkExportItem:
    """Parsed document bundle used to build the bulk master workbook."""

    input_path: Path
    parsed_pages: list[ParsedDocument]
    export_result: ExportResult


class DocumentExporter:
    """Export parsed OCR documents to formatted Excel and JSON files."""

    def export_document(
        self,
        input_path: Path,
        parsed_pages: list[ParsedDocument],
        output_dir: Path,
    ) -> ExportResult:
        """Export one input document into a workbook and JSON file."""
        resolved_output = Path(output_dir).expanduser().resolve()
        resolved_output.mkdir(parents=True, exist_ok=True)

        timestamp = self._timestamp_for_filename()
        safe_stem = self._safe_filename(Path(input_path).stem)
        excel_path = resolved_output / f"{safe_stem}{OUTPUT_STEM_SUFFIX}_{timestamp}{EXCEL_SUFFIX}"
        json_path = resolved_output / f"{safe_stem}{OUTPUT_STEM_SUFFIX}_{timestamp}{JSON_SUFFIX}"

        metadata = self._metadata(input_path=input_path, parsed_pages=parsed_pages)
        workbook = self._build_workbook(input_path=input_path, parsed_pages=parsed_pages, metadata=metadata)
        workbook.save(excel_path)

        self._export_json(json_path=json_path, metadata=metadata, parsed_pages=parsed_pages)

        return ExportResult(
            input_file=Path(input_path).name,
            excel_path=excel_path,
            json_path=json_path,
            table_count=metadata["tables_found"],
            key_value_count=metadata["key_value_pairs_found"],
            page_count=metadata["page_count"],
        )

    def export_bulk_master(self, items: list[BulkExportItem], output_dir: Path) -> Path:
        """Create one master workbook with a summary sheet and one sheet per file."""
        resolved_output = Path(output_dir).expanduser().resolve()
        resolved_output.mkdir(parents=True, exist_ok=True)

        timestamp = self._timestamp_for_filename()
        master_path = resolved_output / f"{MASTER_WORKBOOK_PREFIX}_{timestamp}{EXCEL_SUFFIX}"

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = MASTER_SUMMARY_SHEET_NAME
        self._write_master_summary(summary_sheet, items)

        used_names = {MASTER_SUMMARY_SHEET_NAME}
        for item in items:
            sheet_name = self._unique_sheet_name(self._safe_sheet_name(item.input_path.stem), used_names)
            used_names.add(sheet_name)
            worksheet = workbook.create_sheet(sheet_name)
            self._write_file_sheet(worksheet, item)

        for worksheet in workbook.worksheets:
            self._format_worksheet(worksheet)

        workbook.save(master_path)
        return master_path

    def _build_workbook(
        self,
        input_path: Path,
        parsed_pages: list[ParsedDocument],
        metadata: dict[str, Any],
    ) -> Workbook:
        """Create an openpyxl workbook for one parsed input file."""
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = SUMMARY_SHEET_NAME
        self._write_summary_sheet(summary_sheet, metadata)

        used_names = {SUMMARY_SHEET_NAME}
        table_counter = 0
        for page in parsed_pages:
            for table_index, dataframe in enumerate(page.tables, start=1):
                table_counter += 1
                base_name = f"Page_{page.page_number}_Table_{table_index}"
                sheet_name = self._unique_sheet_name(base_name, used_names)
                used_names.add(sheet_name)
                worksheet = workbook.create_sheet(sheet_name)
                self._write_dataframe(worksheet, dataframe)

        if table_counter == 0:
            worksheet = workbook.create_sheet(NO_TABLES_SHEET_NAME)
            worksheet.append(["Note"])
            worksheet.append([f"No tables were detected in {Path(input_path).name}."])

        key_values_sheet = workbook.create_sheet(KEY_VALUES_SHEET_NAME)
        self._write_key_values_sheet(key_values_sheet, parsed_pages)

        for worksheet in workbook.worksheets:
            self._format_worksheet(worksheet)

        return workbook

    def _write_summary_sheet(self, worksheet: Worksheet, metadata: dict[str, Any]) -> None:
        """Write the Summary worksheet for an individual export."""
        worksheet.append(["Field", "Value"])
        summary_rows = [
            ("File Name", metadata["file_name"]),
            ("Page Count", metadata["page_count"]),
            ("Extraction Date", metadata["extraction_date"]),
            ("Tables Found", metadata["tables_found"]),
            ("Key-Value Pairs Found", metadata["key_value_pairs_found"]),
        ]
        for label, value in summary_rows:
            worksheet.append([label, value])

    def _write_master_summary(self, worksheet: Worksheet, items: list[BulkExportItem]) -> None:
        """Write the summary sheet for a bulk master workbook."""
        worksheet.append(["File", "Pages", "Tables", "Key-Value Pairs", "Excel Output", "JSON Output"])
        for item in items:
            worksheet.append(
                [
                    item.input_path.name,
                    item.export_result.page_count,
                    item.export_result.table_count,
                    item.export_result.key_value_count,
                    str(item.export_result.excel_path),
                    str(item.export_result.json_path),
                ]
            )

    def _write_file_sheet(self, worksheet: Worksheet, item: BulkExportItem) -> None:
        """Write a compact per-file sheet for the bulk master workbook."""
        metadata = self._metadata(input_path=item.input_path, parsed_pages=item.parsed_pages)

        worksheet.append(["Field", "Value"])
        worksheet.append(["File Name", metadata["file_name"]])
        worksheet.append(["Page Count", metadata["page_count"]])
        worksheet.append(["Tables Found", metadata["tables_found"]])
        worksheet.append(["Key-Value Pairs Found", metadata["key_value_pairs_found"]])
        worksheet.append([])
        worksheet.append(["Key", "Value"])

        combined_key_values = self._combined_key_values(item.parsed_pages)
        if combined_key_values:
            for key, value in combined_key_values.items():
                worksheet.append([key, value])
        else:
            worksheet.append(["No key-value pairs detected", ""])

        for page in item.parsed_pages:
            for table_index, dataframe in enumerate(page.tables, start=1):
                worksheet.append([])
                worksheet.append([f"Page {page.page_number} Table {table_index}"])
                start_row = worksheet.max_row + 1
                self._append_dataframe_rows(worksheet, dataframe)
                self._style_header_row(worksheet, start_row)

    def _write_dataframe(self, worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
        """Write a DataFrame into a worksheet."""
        if dataframe.empty:
            worksheet.append(["Note"])
            worksheet.append(["No rows detected"])
            return
        self._append_dataframe_rows(worksheet, dataframe)

    def _append_dataframe_rows(self, worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
        """Append DataFrame header and data rows to a worksheet."""
        for row in dataframe_to_rows(dataframe, index=False, header=True):
            worksheet.append([self._excel_safe_value(value) for value in row])

    def _write_key_values_sheet(self, worksheet: Worksheet, parsed_pages: list[ParsedDocument]) -> None:
        """Write key-value pairs into the final worksheet."""
        worksheet.append(["Key", "Value"])
        combined_key_values = self._combined_key_values(parsed_pages)

        if not combined_key_values:
            worksheet.append(["No key-value pairs detected", ""])
            return

        for key, value in combined_key_values.items():
            worksheet.append([key, value])

    def _combined_key_values(self, parsed_pages: list[ParsedDocument]) -> dict[str, str]:
        """Merge page-level key-value pairs while preserving page context."""
        combined: dict[str, str] = {}
        multi_page = len(parsed_pages) > 1

        for page in parsed_pages:
            for key, value in page.key_values.items():
                output_key = f"Page {page.page_number} - {key}" if multi_page else key
                combined[output_key] = value

        return combined

    def _format_worksheet(self, worksheet: Worksheet) -> None:
        """Apply production-friendly formatting to a worksheet."""
        worksheet.freeze_panes = "A2"
        self._style_header_row(worksheet, 1)
        self._style_section_rows(worksheet)
        self._shade_alternate_rows(worksheet)
        self._auto_width_columns(worksheet)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _style_header_row(self, worksheet: Worksheet, row_number: int) -> None:
        """Apply header styling to one worksheet row."""
        for cell in worksheet[row_number]:
            if cell.value is None:
                continue
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
            cell.font = Font(bold=True, color=HEADER_FONT)

    def _style_section_rows(self, worksheet: Worksheet) -> None:
        """Style single-cell section title rows inside bulk sheets."""
        for row in worksheet.iter_rows(min_row=2):
            values = [cell.value for cell in row if cell.value not in (None, "")]
            if len(values) == 1 and str(values[0]).lower().startswith("page "):
                for cell in row:
                    cell.fill = PatternFill(fill_type="solid", fgColor=SECTION_FILL)
                    cell.font = Font(bold=True)

    def _shade_alternate_rows(self, worksheet: Worksheet) -> None:
        """Apply alternating row shading below the header row."""
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row_index % 2 != 0:
                continue
            for cell in row:
                if cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb not in {None, "00000000", WHITE_FILL}:
                    continue
                cell.fill = PatternFill(fill_type="solid", fgColor=ALTERNATE_ROW_FILL)

    def _auto_width_columns(self, worksheet: Worksheet) -> None:
        """Resize columns based on their visible content."""
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(MAX_COLUMN_WIDTH, max(MIN_COLUMN_WIDTH, max_length + 2))
            worksheet.column_dimensions[column_letter].width = adjusted_width or DEFAULT_COLUMN_WIDTH

    def _export_json(self, json_path: Path, metadata: dict[str, Any], parsed_pages: list[ParsedDocument]) -> None:
        """Write all parsed document data to JSON."""
        payload = {
            "metadata": metadata,
            "pages": [page.to_dict() for page in parsed_pages],
        }
        json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    def _metadata(self, input_path: Path, parsed_pages: list[ParsedDocument]) -> dict[str, Any]:
        """Build export metadata for summaries and JSON."""
        return {
            "file_name": Path(input_path).name,
            "page_count": len(parsed_pages),
            "extraction_date": datetime.now().astimezone().isoformat(timespec="seconds"),
            "tables_found": sum(len(page.tables) for page in parsed_pages),
            "key_value_pairs_found": sum(len(page.key_values) for page in parsed_pages),
        }

    def _safe_sheet_name(self, value: str) -> str:
        """Sanitize a string for use as an Excel sheet name."""
        cleaned = INVALID_SHEET_CHARS.sub("_", value).strip() or "Sheet"
        return cleaned[:MAX_SHEET_NAME_LENGTH]

    def _unique_sheet_name(self, base_name: str, used_names: set[str]) -> str:
        """Return a unique Excel sheet name within the 31-character limit."""
        cleaned_base = self._safe_sheet_name(base_name)
        if cleaned_base not in used_names:
            return cleaned_base

        suffix = 2
        while True:
            suffix_text = f"_{suffix}"
            candidate = f"{cleaned_base[: MAX_SHEET_NAME_LENGTH - len(suffix_text)]}{suffix_text}"
            if candidate not in used_names:
                return candidate
            suffix += 1

    def _safe_filename(self, value: str) -> str:
        """Sanitize a filename stem for generated exports."""
        cleaned = INVALID_FILENAME_CHARS.sub("_", value).strip("._")
        return cleaned or "document"

    def _timestamp_for_filename(self) -> str:
        """Return a filesystem-safe timestamp."""
        return datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")

    def _excel_safe_value(self, value: Any) -> Any:
        """Convert pandas missing values into Excel-friendly blanks."""
        if pd.isna(value):
            return ""
        return value
